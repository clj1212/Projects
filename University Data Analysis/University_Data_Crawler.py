#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Thu Nov  8 21:01:25 2018

@author: zhengmeihua
"""

from selenium import webdriver
import time
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
import pandas as pd
import os
os.chdir('D:/BA/Homework/HW02')

driver = webdriver.Firefox(); 
driver.get("https://souky.eol.cn/api/newapi/assess_result");

driver.find_element_by_css_selector("div.eval-item:nth-child(2)").click(); #点击页面上的'按高校查询'
time.sleep(10) #等待十秒直到页面完全加载出来

wb = Workbook() #新建一个工作簿
ws1 = wb.active
wb.remove(ws1) #删掉初始自带的第一张SHEET

wb2 = load_workbook(filename = 'Universities.xlsx', data_only=True)  #读取Universities.xlsx文件
ws2 = wb2.active
UniName=list(ws2.values)  #读取excel文件中第一张SHEET中的数据并将其转化为列表
UniNameList=[]
for i in range(len(UniName)):
    UniNameList.append(UniName[i][0])  #因为UniName中每个大学名称外侧有层小括号，此步骤是把小括号去掉，是列表的每个元素都是大学名称的字符串

for UniName in UniNameList:   #设置一个循环，把UniNameList中的每个大学搜索一遍，并提取数据到新的SHEET中
    inputbox = driver.find_element_by_css_selector("#sousuoname") 
    inputbox.clear() #清空搜索框
    inputbox.send_keys(UniName)  #填写搜索框
    submitbutton= driver.find_element_by_css_selector("#sousuo")
    submitbutton.click()  #点击‘查询’
    time.sleep(2)
    
    elems = driver.find_element_by_xpath("/html/body/div[5]/div[2]/div/table/tbody/tr[1]") #提取第一行表头信息
    ColumnNames=elems.text.replace('\n',' ').split(" ") #提取标签内的数据并把这些数据设置成列表ColumnNames
    df = pd.DataFrame(columns = ColumnNames)  #把从网页中提取出来的列名赋予新的数据框df
  

    elem1 = driver.find_element_by_xpath("/html/body/div[5]/div[2]/div/table/tbody/tr[1]/td[2]")   #提取‘学校代码’这几个字
    elem2 = driver.find_element_by_xpath("/html/body/div[5]/div[2]/div/table/tbody/tr[2]/td[2]")   #提取这个学校的学校代码
    elem3 = driver.find_element_by_xpath("/html/body/div[5]/div[2]/div/table/tbody/tr[1]/td[3]")   #提取’学校名称'这几个字
    elem4 = driver.find_element_by_xpath("/html/body/div[5]/div[2]/div/table/tbody/tr[2]/td[3]")   #提取这个学校的学校名称
    row1 = [elem1.text,elem2.text,elem3.text,elem4.text]   #把上述几个数据做成列表row1

    ws = wb.create_sheet(title = UniName)   #新建一张SHEET,并命名为这个大学的学校名称
    ws.append(row1) #把row1的数据加在第一行

    elems1 = driver.find_elements_by_xpath("/html/body/div[5]/div[2]/div/table/tbody/tr")  #提取网页上表格内所有的信息
    
    idx=0
    for elem in elems1:  #设置一个循环，把每行数据依次加到df数据框中
        rowdata = elem.text.replace('\n',' ').split(" ")   #提取每行信息每个标签内的数据并设置成列表rowdata
        df.loc[idx]=rowdata  #把这行数据加到df数据框中
        idx += 1 
    df = df[1:]  #去掉第0行，也就是网页上的表头信息，已经将其设置成列名称，所以将其去掉
    df.drop(['学校代码','学校名称'],axis=1, inplace=True)  #删掉df中‘学校代码’和‘学校名称’两列

    for row in dataframe_to_rows(df, index=False, header=True): #将df中的数据写入以学校名称命名的SHEET中
        ws.append(row)
    ws.column_dimensions['C'].width = 22  #设置C列列宽
    ws.column_dimensions['B'].width = 14  #设置B列列宽

wb.save("Univ_Search_Results.xlsx")   #保存文件
wb.close()   #关闭指针
wb2.close()