#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Fri Nov  9 01:07:42 2018

@author: cuilinjing
"""

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font
import pandas as pd
import os
os.chdir('D:/BA/Homework/HW02') 

wb = load_workbook(filename = 'Univ_Eval_AllData.xlsx', data_only = True)  #读取Univ_Eval_AllData.xlsx文件
ws = wb.active

AllData = list(ws.values)  
ColumnsName = list(AllData[0])
data = []
for i in range (1,len(AllData)):
    data.append(list(AllData[i]))   
df = pd.DataFrame(data,columns=ColumnsName)  #上述这一段是把文件中的列名称提取出来，做成新建数据框df的列名称，把文件中列名称下方的数据提取出来加到数据框df中
  
ScoreIndex={'A+':100, 'A': 96, 'A-':92, 'B+':88, 'B':84, 'B-':80, 'C+':75, 'C':70, 'C-':65, 'D':60, 'D-':55, 'F':0}   #设置一个字典，给每个等级设置一个对应的分数
df['分数'] = df['评选结果'].map(ScoreIndex)   #给df新建一列‘分数’，其中是根据ScoreIndex将每个评选结果对应的分数加在这新的一列上

SubjectList=list(df['一级学科代码'].values)  #将df中‘一级学科代码’列中的所有数据提取并转换成列表SubjectList
CategoryList = []
for i in range(len(SubjectList)):  #根据SubjectList中每个元素的前两个字符提取出来转换成整数型数值，并根据这个数值判断其对应的专业大类，将每个学科对应的专业大类设置成列表CategoryList
    CategoryNum = int(SubjectList[i][0:2])
    if CategoryNum in [1,5,13]:
        Category = '人文科学'
    if CategoryNum in [2,3,4,6,11,12]:
        Category = '社会科学'
    if CategoryNum == 7:
        Category = '自然科学'
    if CategoryNum == 8:
        Category = '工程与技术科学'
    if CategoryNum == 10:
        Category = '医药科学'
    if CategoryNum == 9:
        Category = '农业科学'
    CategoryList.append(Category)
df['专业大类'] = CategoryList  #给df新建一列'专业大类',把CategoryList中的数据加在这列下方

df = df.sort_values(by=['专业大类','一级学科代码','分数'],ascending=[True,True,False])  #将df先按照“专业大类”升序排序,再按“一级学科代码”升序排序,最后按“分数”降序排序

wb = Workbook()  #新建一个工作簿
ws = wb.active  
ws.title = 'MajorRanks'  #将初始自带的SHEET重命名为MajorRanks

for row in dataframe_to_rows(df, index=False, header=True):   #把数据框df中的数据写入MajorRanks
    ws.append(row)

ws.column_dimensions['B'].width=16   #设置B列列宽
ws.column_dimensions['C'].width=12   #设置C列列宽
ws.column_dimensions['D'].width=19   #设置D列列宽
ws.column_dimensions['G'].width=15   #设置G列列宽

wb.save('Univ_Eval_MajorRanks.xlsx')   #保存文件，并命名为Univ_Eval_MajorRanks.xlsx

MyPatterns = [PatternFill(fgColor='EEFFCC',fill_type='solid'),PatternFill(fgColor='E6EEFF',fill_type='solid')]   #定义背景的填色

initSubject = ws['D2'].value   #读取D列第二行的数据（即MajorRanks出现的第一个一级学科名称）赋值给initSubject
rowN = 1
a = 1
for row in ws.iter_rows(min_row=2, max_row = ws.max_row, max_col = ws.max_column): 
    #设置一个循环，提取每行B列和D列的数据，如果B列数据为‘上海交通大学’，则将此行的每个单元格字体颜色定义为红色
    #如果D列数据与上一行D列数据相同，就把此行每个单元格背景颜色定义为与上一行相同的背景颜色
    #如果D列数据与上一行D列数据不同，就把此行每个单元格颜色定义为与上一行不同的另一种背景颜色
    rowN += 1
    school = ws['B'+str(rowN)].value   
    newSubject = ws['D'+str(rowN)].value
    if school =='上海交通大学':
        for cell in row:
            cell.font = Font(color='FF0000')
    if newSubject == initSubject:
        ColIdx = a%2
        for cell in row:
            cell.fill = MyPatterns[ColIdx]
    else:
        a += 1
        ColIdx = a%2
        for cell in row:
            cell.fill = MyPatterns[ColIdx]
    initSubject = newSubject
    
    
wb.save('Univ_Eval_MajorRanks.xlsx')   #保存文件
wb.close()   #关闭指针

