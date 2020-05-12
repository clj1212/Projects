# -*- coding: utf-8 -*-
"""
Created on Thu Nov  8 23:11:13 2018

@author: 汪雅男
"""
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
import os

os.chdir('D:/BA/Homework/HW02')
df1 = pd.read_excel('D:/BA/Homework/HW02/Univ_Eval_MajorRanks.xlsx', encoding='UTF-8')
MyAgg = {'一级学科代码': 'count'}#使用MyAgg根据一级学科代码来进行分类汇总
pvtb = df1.groupby(['学校名称','专业大类']).agg(MyAgg)#建立双层分组索引，并将各学科数加总填入
pvtb = pd.DataFrame(pvtb)#把结果转换成数据框使用unstack函数取消堆叠
pvtb = pvtb.unstack()#使用unstack函数取消堆叠
Cols = pvtb.columns.levels[1] #提取有用的列名转换成list格式
ColNames = pd.Series(Cols, name ='专业大类')#提取出键值为专业大类的这一列，并赋值给ColNames
pvtb.columns = ColNames#修改列名为以专业大类为键值的这一列
pvtb = pvtb.fillna(0)#将数据框中所有的NA换成0
pvtb = pvtb.astype(int)#将浮点数转为整数

pvtb['总计'] = pvtb.apply(sum,axis=1)#加总各校学科数
pvtb.sort_values(by='总计',ascending=False,inplace=True)#按总计的降序排列


    
    
df2=df1[df1.分数>=92]#筛选出分数大于等于92的专业
pvtb1 = df2.groupby(['学校名称','专业大类']).agg(MyAgg)#与df1同理，建立双层分组索引，并将学科数加总填入
pvtb1 = pd.DataFrame(pvtb1)#把结果转换成数据框
pvtb1 = pvtb1.unstack()#使用unstack函数取消堆叠
Cols1 = pvtb1.columns.levels[1] #提取有用的列名转换成list格式
ColNames1 = pd.Series(Cols1, name ='专业大类')#提取出键值为专业大类的这一列，并赋值给ColNames1
pvtb1.columns = ColNames1#修改列名为以专业大类为键值的这一列
pvtb1 = pvtb1.fillna(0)#将数据框中所有的NA换成0
pvtb1 = pvtb1.astype(int)#将浮点数转为整数

pvtb1['总计'] = pvtb1.apply(sum,axis=1)#加总各专业数
pvtb1.sort_values(by='总计',ascending=False,inplace=True)#按总计的降序排列




df3=df1[df1.分数>=96]#筛选出分数大于等于96的专业
pvtb2 = df3.groupby(['学校名称','专业大类']).agg(MyAgg)#与df1同理，建立双层分组索引，并将学科数加总填入
pvtb2 = pd.DataFrame(pvtb2)#把结果转换成数据框
pvtb2 = pvtb2.unstack()#使用unstack函数取消堆叠
Cols2 = pvtb2.columns.levels[1] #提取有用的列名转换成list格式
ColNames2 = pd.Series(Cols2, name ='专业大类')#提取出键值为专业大类的这一列，并赋值给ColNames2
pvtb2.columns = ColNames2#修改列名为以专业大类为键值的这一列
pvtb2 = pvtb2.fillna(0)#将数据框中所有的NA换成0
pvtb2 = pvtb2.astype(int)#将浮点数转为整数

pvtb2['总计'] = pvtb2.apply(sum,axis=1)#加总各专业数
pvtb2.sort_values(by='总计',ascending=False,inplace=True)#按总计的降序排列



wb = load_workbook(filename='Univ_Eval_MajorRanks.xlsx',data_only=True)

ws1 = wb.create_sheet('MajorsEval')#建立新的worksheet，命名为MajorsEval
ws1.column_dimensions['A'].width = 17#设置列A列宽为17
for row in dataframe_to_rows(pvtb, index=True, header=True): #遍历pvtb，填充数据
    ws1.append(row)
ws1['A1']='专业大类'#将A1取值为'专业大类'


ws2 = wb.create_sheet('A_Majors1')#建立新的worksheet，命名为A_Majors1
ws2.column_dimensions['A'].width = 17#设置列A列宽为17
for row in dataframe_to_rows(pvtb1, index=True, header=True):#遍历pvtb1，填充数据
    ws2.append(row)
ws2['A1']='专业大类'#将A1取值为'专业大类'

ws3 = wb.create_sheet('A_Majors2')#建立新的worksheet，命名为A_Majors2
ws3.column_dimensions['A'].width = 17#设置列A列宽为17
for row in dataframe_to_rows(pvtb2, index=True, header=True):#遍历pvtb2，填充数据
    ws3.append(row)
ws3['A1']='专业大类'#将A1取值为'专业大类'
wb.save("Univ_Eval_MajorRanks.xlsx")
wb.close()

