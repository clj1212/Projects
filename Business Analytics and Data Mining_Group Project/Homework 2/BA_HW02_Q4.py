# -*- coding: utf-8-*-
"""
Created on Sun Aug 12 13:46:12 2018
@author: Zach Zhizhong ZHOU 周志中
@institution: Shanghai Jiao Tong University 上海交通大学
"""
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
import openpyxl.utils.cell as cellutl
import pandas as pd
import os
os.chdir('/Users/zhengmeihua/Documents/大三上/商业统计分析与数据挖掘/Homework/HW02')
import Get_Excel_Range as getExcR

wb = load_workbook(filename = 'Univ_Search_Results.xlsx')
Sheets = wb.sheetnames

new_wb= Workbook() #这种新建文件方式会自动生成一个名为Sheet的工作簿。
new_ws = new_wb['Sheet']

tb_head = ['学校代码','学校名称','一级学科代码','一级学科名称','评选结果']
#加上注释说明下面一行语句做了什么事情。
    #用append函数往名为sheet的工作簿添加一行数据，A1取值为‘学校代码’，B1取值为‘学校名称’，C1取值为‘一级学科代码’，D1取值为‘一级学科名称’，E1取值为‘评选成果’
new_ws.append(tb_head)


for sheetname in Sheets:
    ws = wb[sheetname]
#加上注释说明下面一行语句做了什么事情。
    #此for循环每遍历一次，max_row就被赋值为当前worksheet的最大行数（有效行数）
    max_row = cellutl.range_boundaries(ws.dimensions)[3]#max row number
#加上注释说明下面一行语句做了什么事情。
    #将BCD这三列从第三行到最后一行（最大行数），这一区块的数据赋值给tb
    tb = getExcR.load_workbook_range('B3:D'+str(max_row),ws)
    print(tb)
#加上注释说明下面2行语句做了什么事情。
    #以B1和D1的取值各自生成行数为最大行数减二（max_row-2）的两列，以行索引合并之后再与之前生成的td合并
    tb = pd.concat([pd.Series([ws['B1'].value]*(max_row-2)), \
                  pd.Series([ws['D1'].value]*(max_row-2)), tb ], axis=1)
#加上注释说明下面2行语句做了什么事情。
    #通过行（row）遍历整个工作表填充相应的数据，并且工作表的行index和header都被舍弃
    for row in dataframe_to_rows(tb, index=False, header=False):
        new_ws.append(row)
  

#提供两种存储数据方法，第一种：存成csv文件，编码选择GB2312
#AllData.columns = tb_head
#AllData.to_csv("Univ_Eval_AllData.csv",index=False,sep=',',encoding='GB2312')
#存入csv的中文编码默认是utf-8，使用Excel打开时候是乱码。
#如果一定想用Excel看这个文件，可以加上 encoding='GB2312'

new_wb.save('Univ_Eval_AllData.xlsx')
new_wb.close()
wb.close()
